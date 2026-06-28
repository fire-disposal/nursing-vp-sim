"""Generic export engine — format-agnostic, reusable tabular export."""

import csv
import io
from abc import ABC, abstractmethod
from collections.abc import Callable, Sequence
from dataclasses import dataclass
from typing import Any, Generic, TypeVar
from urllib.parse import quote

from fastapi.responses import Response

T = TypeVar("T")


@dataclass
class ColumnDef(Generic[T]):
    """Column: header + key (attribute/dict-key) or value (callable on item) + optional fmt."""

    header: str
    key: str = ""
    value: Callable[[T], Any] | None = None
    fmt: Callable[[Any], str | None] | None = None

    def _resolve(self, item: T) -> Any:
        if self.value is not None:
            return self.value(item)
        return getattr(item, self.key, None)


def _sanitize_csv(val: str | None) -> str:
    if val is None:
        return ""
    if val and val[0] in ("=", "+", "-", "@"):
        return "'" + val
    return val


class Exporter(ABC, Generic[T]):
    """Base export engine."""

    @abstractmethod
    def export(self, items: Sequence[T], columns: list[ColumnDef[T]], title: str = "") -> bytes: ...


class CSVExporter(Exporter[T]):
    """Export to CSV (UTF-8 BOM)."""

    def export(self, items: Sequence[T], columns: list[ColumnDef[T]], title: str = "") -> bytes:
        buf = io.StringIO()
        buf.write("\ufeff")
        w = csv.writer(buf)
        w.writerow([c.header for c in columns])
        for item in items:
            row = []
            for c in columns:
                raw = c._resolve(item)
                if raw is None:
                    row.append("")
                elif c.fmt:
                    row.append(_sanitize_csv(c.fmt(raw)))
                else:
                    row.append(_sanitize_csv(str(raw)))
            w.writerow(row)
        return buf.getvalue().encode("utf-8-sig")


class XLSXExporter(Exporter[T]):
    """Export to .xlsx with styled header row and auto column widths."""

    _HEADER_FILL = "2563EB"

    def __init__(self, column_width: int = 18):
        self._col_width = column_width

    def export(self, items: Sequence[T], columns: list[ColumnDef[T]], title: str = "") -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = (title or "Sheet1")[:31]

        hfont = Font(bold=True, color="FFFFFF")
        hfill = PatternFill(start_color=self._HEADER_FILL, end_color=self._HEADER_FILL, fill_type="solid")
        halign = Alignment(horizontal="center")

        for ci, c in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=c.header)
            cell.font = hfont
            cell.fill = hfill
            cell.alignment = halign

        for ri, item in enumerate(items, 2):
            for ci, c in enumerate(columns, 1):
                raw = c._resolve(item)
                if raw is None:
                    val = ""
                elif c.fmt:
                    val = c.fmt(raw)
                else:
                    val = raw
                ws.cell(row=ri, column=ci, value=val)

        for ci in range(1, len(columns) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = self._col_width

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


def export_response(
    items: Sequence[T], columns: list[ColumnDef[T]], filename: str, title: str = "", format: str = "csv"
) -> Response:
    """Build a FastAPI Response exporting *items* in the given *format* (csv|xlsx)."""
    ext = format
    encoded = quote(filename)
    disposition = f"attachment; filename*=UTF-8''{encoded}.{ext}"

    if format == "xlsx":
        content = XLSXExporter().export(items, columns, title or filename)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        content = CSVExporter().export(items, columns, title or filename)
        media_type = "text/csv; charset=utf-8-sig"
    return Response(content=content, media_type=media_type, headers={"Content-Disposition": disposition})
