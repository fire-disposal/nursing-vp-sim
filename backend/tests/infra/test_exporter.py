"""Unit tests for the generic export engine (CSV / XLSX / response)."""

import csv
import io
from types import SimpleNamespace

import pytest
from fastapi import HTTPException

from infra.exporter import ColumnDef, CSVExporter, XLSXExporter, export_response


def _records():
    return [
        SimpleNamespace(name="张三", age=30, score=91.5),
        SimpleNamespace(name=None, age=0, score=None),
    ]


def _columns():
    return [
        ColumnDef("姓名", key="name"),
        ColumnDef("年龄", key="age"),
        ColumnDef("成绩", key="score", fmt=lambda v: f"{v:.1f}"),
        ColumnDef("问候", value=lambda r: f"你好{r.name or ''}"),
    ]


class TestColumnDef:
    def test_resolve_by_key(self):
        col = ColumnDef("年龄", key="age")
        assert col._resolve(SimpleNamespace(age=5)) == 5

    def test_resolve_by_callable(self):
        col = ColumnDef("翻倍", value=lambda r: r.age * 2)
        assert col._resolve(SimpleNamespace(age=5)) == 10

    def test_missing_key_returns_none(self):
        col = ColumnDef("缺失", key="nope")
        assert col._resolve(SimpleNamespace(age=5)) is None


class TestCSVExporter:
    def test_single_utf8_bom_only(self):
        data = CSVExporter().export(_records(), _columns())
        # 恰好一个 BOM（utf-8-sig 编码产生），不存在双 BOM
        assert data.count(b"\xef\xbb\xbf") == 1
        assert data.startswith(b"\xef\xbb\xbf")

    def test_rows_and_format_applied(self):
        data = CSVExporter().export(_records(), _columns())
        text = data.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        assert rows[0] == ["姓名", "年龄", "成绩", "问候"]
        assert rows[1][:3] == ["张三", "30", "91.5"]
        assert rows[1][3] == "你好张三"
        # None → empty string
        assert rows[2][0] == ""
        assert rows[2][2] == ""

    def test_formula_injection_sanitized(self):
        cols = [ColumnDef("值", key="v")]
        records = [SimpleNamespace(v="=cmd()"), SimpleNamespace(v="+1"), SimpleNamespace(v="safe")]
        data = CSVExporter().export(records, cols)
        text = data.decode("utf-8-sig")
        rows = list(csv.reader(io.StringIO(text)))
        assert rows[1][0] == "'=cmd()"
        assert rows[2][0] == "'+1"
        assert rows[3][0] == "safe"


class TestXLSXExporter:
    def test_produces_valid_workbook(self):
        from openpyxl import load_workbook

        data = XLSXExporter().export(_records(), _columns(), title="测试导出")
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        assert ws.title == "测试导出"
        assert ws.cell(row=1, column=1).value == "姓名"
        assert ws.cell(row=2, column=1).value == "张三"
        assert ws.cell(row=2, column=3).value == "91.5"
        assert ws.cell(row=3, column=2).value == 0

    def test_title_truncated_to_31_chars(self):
        from openpyxl import load_workbook

        long_title = "长" * 40
        data = XLSXExporter().export([], [ColumnDef("h", key="x")], title=long_title)
        wb = load_workbook(io.BytesIO(data))
        assert len(wb.active.title) == 31


class TestExportResponse:
    def test_csv_response(self):
        from urllib.parse import unquote

        resp = export_response(_records(), _columns(), "成绩单")
        assert resp.media_type == "text/csv; charset=utf-8-sig"
        disposition = resp.headers["Content-Disposition"]
        assert "attachment" in disposition
        assert "成绩单" in unquote(disposition)

    def test_xlsx_response(self):
        resp = export_response(_records(), _columns(), "成绩单", format="xlsx")
        assert resp.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert resp.body.startswith(b"PK")  # zip magic

    def test_rejects_too_many_rows(self, monkeypatch):
        from core import config

        monkeypatch.setattr(config, "MAX_EXPORT_ROWS", 1)
        with pytest.raises(HTTPException) as exc:
            export_response(_records(), _columns(), "成绩单")
        assert exc.value.status_code == 400
